#!/usr/bin/env python3
"""Phase 1 of /expense-report — collect.

Sweep Gmail in a date range for receipt/invoice originals. Parse vendor,
amount, currency, date. Dedupe against existing Odoo records by receipt
number. Write an XLSX with a blank `Include` column for human curation.

Usage:
    python collect.py <start_yyyy-mm-dd> <end_yyyy-mm-dd> [out_xlsx_path]
"""
from __future__ import annotations

import json
import re
import sys
from datetime import datetime
from pathlib import Path

# Allow running directly from the skill scripts/ dir or via absolute path
sys.path.insert(0, str(Path(__file__).parent))
import _lib as L  # noqa: E402

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402


# Gmail search queries — broad enough to catch SaaS receipts, vendor
# notifications, and forwarded copies; we filter Fwd:/Re: in post-processing.
def queries(start: str, end: str) -> list[str]:
    s, e = start.replace("-", "/"), end.replace("-", "/")
    return [
        # SaaS receipts directly from vendors
        f"(from:anthropic.com OR from:warp.dev OR from:vercel.com OR "
        f"from:openrouter.ai OR from:wisprflow.ai OR from:gladia.io OR "
        f"from:elevenlabs.io OR from:xai.com OR from:x.ai OR from:openai.com OR "
        f"from:mainfunc.ai OR from:agentmail.to OR from:odoo.com) "
        f"(receipt OR invoice OR subscription) after:{s} before:{e}",
        # Travel
        f"(from:booking.com OR from:ryanair.com OR from:lufthansa.com OR "
        f"from:ita-airways.com OR from:trenitalia.com OR from:italo.it OR "
        f"from:uber.com OR from:freenow.com OR from:airbnb.com) "
        f"(receipt OR confirmation OR ticket OR itinerary) "
        f"after:{s} before:{e}",
        # Anything explicitly tagged as a receipt/invoice
        f"(subject:receipt OR subject:invoice OR subject:fattura OR "
        f"subject:ricevuta OR subject:scontrino) "
        f"after:{s} before:{e} -in:trash -from:david.orban@gmail.com",
        # Forwards (we'll filter, but useful as fallback)
        f"to:expenses@expensify.com after:{s} before:{e}",
    ]


SKIP_SUBJECT_RE = [
    re.compile(r"^Fwd:", re.I),
    re.compile(r"^Re:", re.I),
    re.compile(r"unsuccessful", re.I),
    re.compile(r"past due", re.I),
    re.compile(r"action needed", re.I),
    re.compile(r"^you have a message", re.I),
    re.compile(r"impact begins", re.I),
]


def main(start: str, end: str, out_path: str | None) -> None:
    # Validate dates
    try:
        datetime.strptime(start, "%Y-%m-%d")
        datetime.strptime(end, "%Y-%m-%d")
    except ValueError:
        print("Dates must be YYYY-MM-DD", file=sys.stderr)
        sys.exit(2)

    if not out_path:
        out_path = f"/Users/davidorban/Downloads/Futuroid_Expenses_{start}_to_{end}.xlsx"
    out = Path(out_path)

    print("Auth Gmail...", file=sys.stderr)
    token = L.gmail_token()

    existing_nums = L.get_existing_receipt_numbers()
    print(f"Existing Odoo receipt #s: {len(existing_nums)}", file=sys.stderr)

    msg_ids: dict[str, None] = {}
    for q in queries(start, end):
        print(f"Query: {q[:100]}...", file=sys.stderr)
        for m in L.list_messages(token, q):
            msg_ids.setdefault(m["id"], None)
    print(f"Unique candidates: {len(msg_ids)}", file=sys.stderr)

    parsed: list[dict] = []
    for i, mid in enumerate(msg_ids, 1):
        try:
            full = L.gmail_get(
                f"https://gmail.googleapis.com/gmail/v1/users/me/messages/{mid}?format=full",
                token,
            )
        except Exception as ex:
            print(f"  ! {mid}: {ex}", file=sys.stderr)
            continue
        headers = full.get("payload", {}).get("headers", [])
        subj = next((h["value"] for h in headers if h["name"].lower() == "subject"), "")
        sender = next((h["value"] for h in headers if h["name"].lower() == "from"), "")
        body = L.extract_text(full.get("payload", {}))
        amount, currency = L.parse_amount(body)
        parsed.append({
            "msg_id": mid,
            "date": L.parse_email_date(headers),
            "subject": subj,
            "from": sender,
            "vendor": L.parse_vendor(subj, sender),
            "amount": amount,
            "currency": currency,
            "receipt_no": L.receipt_number(subj),
            "snippet": (body[:200] or "").replace("\n", " ").strip(),
        })
        if i % 20 == 0:
            print(f"  parsed {i}/{len(msg_ids)}", file=sys.stderr)

    # Filter + dedupe
    kept: list[dict] = []
    dropped = dict.fromkeys(
        ["range", "fwd_or_re", "noise", "no_amount", "no_vendor",
         "dup_in_odoo", "dup_self"], 0)
    seen: set[tuple] = set()
    for p in parsed:
        d = p.get("date") or ""
        if d < start or d > end:
            dropped["range"] += 1
            continue
        subj = p.get("subject") or ""
        if any(rx.search(subj) for rx in SKIP_SUBJECT_RE[:2]):
            dropped["fwd_or_re"] += 1
            continue
        if any(rx.search(subj) for rx in SKIP_SUBJECT_RE[2:]):
            dropped["noise"] += 1
            continue
        if not p.get("amount") or not p.get("currency"):
            dropped["no_amount"] += 1
            continue
        if not p.get("vendor"):
            dropped["no_vendor"] += 1
            continue
        if p["amount"] < 0.5:
            dropped["noise"] += 1
            continue
        if p["receipt_no"] and p["receipt_no"] in existing_nums:
            dropped["dup_in_odoo"] += 1
            continue
        key = (p["vendor"], p["receipt_no"] or "", round(p["amount"], 2), d)
        if key in seen:
            dropped["dup_self"] += 1
            continue
        seen.add(key)
        kept.append(p)

    kept.sort(key=lambda x: (x["date"], x["vendor"] or ""))
    print(f"\nKept {len(kept)} | Dropped " + ", ".join(f"{k}={v}" for k, v in dropped.items()),
          file=sys.stderr)

    # Build XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = f"{start} to {end}"
    headers_row = ["Include", "Date", "Vendor", "Receipt #", "Currency",
                   "Amount", "Subject", "From", "Gmail ID", "Snippet"]
    for col, h in enumerate(headers_row, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1C2833")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cur_totals: dict[str, float] = {}
    for i, p in enumerate(kept, 2):
        ws.cell(row=i, column=1, value="")  # blank Include
        ws.cell(row=i, column=2, value=p["date"])
        ws.cell(row=i, column=3, value=p["vendor"])
        ws.cell(row=i, column=4, value=p["receipt_no"] or "")
        ws.cell(row=i, column=5, value=p["currency"])
        amt_cell = ws.cell(row=i, column=6, value=round(p["amount"], 2))
        amt_cell.number_format = "#,##0.00"
        ws.cell(row=i, column=7, value=(p["subject"] or "")[:120])
        ws.cell(row=i, column=8, value=(p["from"] or "")[:60])
        ws.cell(row=i, column=9, value=p["msg_id"])
        ws.cell(row=i, column=10, value=(p["snippet"] or "")[:200])
        cur_totals[p["currency"]] = cur_totals.get(p["currency"], 0) + p["amount"]

    sr = len(kept) + 3
    ws.cell(row=sr, column=1, value=f"TOTALS ({len(kept)} expenses)").font = Font(bold=True)
    for off, (c, t) in enumerate(sorted(cur_totals.items())):
        ws.cell(row=sr + 1 + off, column=5, value=c).font = Font(bold=True)
        cell = ws.cell(row=sr + 1 + off, column=6, value=round(t, 2))
        cell.number_format = "#,##0.00"
        cell.font = Font(bold=True)

    widths = [9, 12, 22, 14, 9, 11, 50, 30, 22, 60]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "B2"

    wb.save(out)
    print(f"\nWrote: {out}", file=sys.stderr)
    print("By currency:", file=sys.stderr)
    for c, t in sorted(cur_totals.items()):
        print(f"  {c}  {t:>10,.2f}", file=sys.stderr)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: collect.py <start_yyyy-mm-dd> <end_yyyy-mm-dd> [out_xlsx]", file=sys.stderr)
        sys.exit(2)
    main(sys.argv[1], sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else None)
    L.safe_exit(0)
