#!/usr/bin/env python3
"""Phase 2 of /expense-report — finalize.

Read a curated XLSX (with `Include` column marked `y` for rows to keep),
push each y row to Odoo as a hr.expense, attach its receipt (PDF or
HTML-rendered) as ir.attachment, and build a merged PDF report.

Usage:
    python finalize.py <xlsx_path> [pdf_out_path]
"""
from __future__ import annotations

import base64
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
import _lib as L  # noqa: E402

from openpyxl import load_workbook  # noqa: E402
from reportlab.lib import colors  # noqa: E402
from reportlab.lib.pagesizes import A4  # noqa: E402
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle  # noqa: E402
from reportlab.lib.units import cm  # noqa: E402
from reportlab.platypus import (  # noqa: E402
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
)
from pypdf import PdfReader, PdfWriter  # noqa: E402

TMPDIR = Path("/tmp/expense-report-skill")
TMPDIR.mkdir(exist_ok=True)


# ----------------- PDF page builders -----------------
STYLES = getSampleStyleSheet()
H1 = ParagraphStyle("h1", parent=STYLES["Heading1"], fontSize=18, spaceAfter=12,
                    textColor=colors.HexColor("#1c2833"))
H2 = ParagraphStyle("h2", parent=STYLES["Heading2"], fontSize=13, spaceAfter=8,
                    textColor=colors.HexColor("#2e4053"))
H3 = ParagraphStyle("h3", parent=STYLES["Heading3"], fontSize=11, spaceAfter=6,
                    textColor=colors.HexColor("#5d6d7e"))
BODY = ParagraphStyle("body", parent=STYLES["BodyText"], fontSize=9.5, leading=12)
SMALL = ParagraphStyle("small", parent=STYLES["BodyText"], fontSize=8, leading=10,
                       textColor=colors.HexColor("#566573"))


def build_cover(rows, usd_rates, rate_date, period_label, out_path):
    doc = SimpleDocTemplate(str(out_path), pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    story = []
    story.append(Paragraph("Futuroid Ltd — Expense Report", H1))
    story.append(Paragraph(
        f"Period: {period_label} · Employee: David Orban · "
        f"Payment mode: Personal (own_account) · Records: {len(rows)}", BODY))
    story.append(Spacer(1, 0.4*cm))

    cur_totals: dict[str, float] = {}
    usd_grand = 0.0
    for r in rows:
        cur_totals[r["currency"]] = cur_totals.get(r["currency"], 0) + r["amount"]
        usd_grand += r["amount"] * usd_rates.get(r["currency"], 0)

    cur_lines = " · ".join(f"<b>{c}</b> {t:,.2f}" for c, t in sorted(cur_totals.items()))
    story.append(Paragraph(f"Totals by currency: {cur_lines}", BODY))
    story.append(Spacer(1, 0.2*cm))
    rate_lines = " · ".join(f"1 {c} = {usd_rates[c]:.4f} USD"
                            for c in sorted(cur_totals) if c in usd_rates)
    story.append(Paragraph(
        f"<b>Grand total: USD {usd_grand:,.2f}</b>",
        ParagraphStyle("gt", parent=BODY, fontSize=12, textColor=colors.HexColor("#1c2833"))
    ))
    story.append(Paragraph(
        f"Conversion rates (ECB via Odoo res.currency.rate, fixed {rate_date}): {rate_lines}",
        SMALL))
    story.append(Spacer(1, 0.6*cm))

    data = [["#", "Date", "Vendor", "Category", "Currency", "Amount"]]
    for i, r in enumerate(rows, 1):
        data.append([str(i), str(r["date"]), Paragraph(r["vendor"], BODY),
                     r["category_label"], r["currency"], f"{r['amount']:,.2f}"])
    tbl = Table(data, colWidths=[1*cm, 2.3*cm, 5.8*cm, 4*cm, 2*cm, 2.5*cm], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1c2833")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("ALIGN", (0, 0), (1, -1), "CENTER"),
        ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#aab7b8")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ]))
    story.append(tbl)
    doc.build(story)


def build_detail(idx, r, has_receipt, out_path):
    doc = SimpleDocTemplate(str(out_path), pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    story = [Paragraph(f"Expense {idx} — Odoo #{r['odoo_id']}", H2),
             Paragraph(r["vendor"], H3)]
    meta = (f"<b>Date:</b> {r['date']} · "
            f"<b>Amount:</b> {r['currency']} {r['amount']:,.2f} · "
            f"<b>Category:</b> {r['category_label']} · "
            f"<b>Gmail ID:</b> {r.get('msg_id') or 'n/a'}")
    story.append(Paragraph(meta, SMALL))
    if r.get("receipt_no"):
        story.append(Paragraph(f"<b>Receipt #:</b> {r['receipt_no']}", BODY))
    if r.get("subject"):
        story.append(Paragraph(f"<b>Subject:</b> {r['subject']}", BODY))
    if has_receipt:
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph("<i>↓ Original receipt follows on next page(s).</i>", BODY))
    doc.build(story)


# ----------------- Main -----------------
def read_y_rows(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}
    if "Include" not in idx:
        raise SystemExit("XLSX has no `Include` column — was this generated by collect.py?")
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        inc = row[idx["Include"]]
        if not inc or str(inc).strip().lower() != "y":
            continue
        d = row[idx["Date"]]
        date_str = d.isoformat() if hasattr(d, "isoformat") else str(d or "")
        if not date_str:
            continue
        rows.append({
            "date": date_str,
            "vendor": row[idx["Vendor"]] or "",
            "receipt_no": row[idx["Receipt #"]] or "",
            "currency": row[idx["Currency"]] or "",
            "amount": float(row[idx["Amount"]] or 0),
            "subject": row[idx["Subject"]] or "",
            "from": row[idx.get("From", -1)] if "From" in idx else "",
            "msg_id": row[idx["Gmail ID"]] or "",
        })
    rows.sort(key=lambda r: (r["date"], r["vendor"]))
    return rows


def push_to_odoo(rows: list[dict], token: str) -> None:
    """Create hr.expense + attach receipt for each row. Mutates rows in place
    with `odoo_id`, `category_id`, `category_label`."""
    db, uid, k, models = L.odoo_session()
    cur_ids = L.get_currency_id_map({"USD", "EUR", "GBP", "RON"})

    for i, r in enumerate(rows, 1):
        pid, label = L.categorize(r["vendor"])
        r["category_id"] = pid
        r["category_label"] = label

        clean = re.sub(r"^Fwd:\s*", "", r["subject"] or "").strip()
        clean = re.sub(r"^Your\s+(?:receipt|invoice)\s+from\s+", "", clean, flags=re.I)
        name = clean[:80] or f"{r['vendor']} {r['currency']} {r['amount']:.2f}"

        vals = {
            "name": name,
            "employee_id": L.DAVID_EMPLOYEE_ID,
            "company_id": L.FUTUROID_COMPANY_ID,
            "date": r["date"],
            "product_id": pid,
            "total_amount_currency": r["amount"],
            "currency_id": cur_ids[r["currency"]],
            "payment_mode": "own_account",
            "description": (
                f"Vendor: {r['vendor']}\n"
                f"Receipt #: {r['receipt_no']}\n"
                f"Gmail ID: {r['msg_id']}\n"
                f"Subject: {r['subject']}"
            ),
        }
        try:
            eid = models.execute_kw(db, uid, k, "hr.expense", "create", [vals])
        except Exception as ex:
            print(f"  {i:2}. ! create failed for {r['vendor']}: {str(ex)[:200]}",
                  file=sys.stderr)
            r["odoo_id"] = None
            continue
        r["odoo_id"] = eid

        # Attach receipt from Gmail
        att_label = "none"
        if r["msg_id"]:
            try:
                full = L.gmail_get(
                    f"https://gmail.googleapis.com/gmail/v1/users/me/messages/{r['msg_id']}?format=full",
                    token)
                payload = full.get("payload", {})
                pdfs = L.find_pdfs(payload)
                if pdfs:
                    fname, att_id = pdfs[0]
                    raw = L.dl_attachment(r["msg_id"], att_id, token)
                    if raw.startswith(b"%PDF"):
                        models.execute_kw(db, uid, k, "ir.attachment", "create",
                            [{"name": fname,
                              "datas": base64.b64encode(raw).decode(),
                              "res_model": "hr.expense", "res_id": eid,
                              "type": "binary", "mimetype": "application/pdf"}])
                        att_label = f"PDF {fname}"
                else:
                    html = L.extract_html(payload)
                    if html:
                        cid_map = L.collect_cid_images(payload, r["msg_id"], token)
                        pdf_bytes = L.html_to_pdf(html, cid_map)
                        if pdf_bytes:
                            models.execute_kw(db, uid, k, "ir.attachment", "create",
                                [{"name": f"{r['vendor']}-receipt.pdf",
                                  "datas": base64.b64encode(pdf_bytes).decode(),
                                  "res_model": "hr.expense", "res_id": eid,
                                  "type": "binary", "mimetype": "application/pdf"}])
                            att_label = "HTML rendered"
            except Exception as ex:
                print(f"     ! attach failed: {str(ex)[:120]}", file=sys.stderr)

        print(f"  {i:2}. #{eid:3} · {r['date']} · {label:25} · "
              f"{r['vendor'][:22]:22} · {r['currency']} {r['amount']:7.2f} · "
              f"receipt: {att_label}", file=sys.stderr)


def fetch_odoo_attachments(expense_id: int) -> list[tuple[str, str, bytes]]:
    db, uid, k, models = L.odoo_session()
    rows = models.execute_kw(
        db, uid, k, "ir.attachment", "search_read",
        [[("res_model", "=", "hr.expense"), ("res_id", "=", expense_id)]],
        {"fields": ["name", "mimetype", "datas"]},
    )
    out = []
    for r in rows:
        data = r.get("datas")
        if not data:
            continue
        try:
            raw = base64.b64decode(data)
        except Exception:
            continue
        out.append((r["name"], r.get("mimetype") or "", raw))
    return out


def build_pdf(rows: list[dict], period_label: str, out_path: Path) -> None:
    usd_rates, rate_date = L.get_live_usd_rates()

    cover = TMPDIR / "00_cover.pdf"
    build_cover(rows, usd_rates, rate_date, period_label, cover)
    pieces: list[Path] = [cover]

    for i, r in enumerate(rows, 1):
        eid = r.get("odoo_id")
        # Pull attachments directly from Odoo (the source of truth)
        atts = fetch_odoo_attachments(eid) if eid else []

        receipt_pdfs: list[Path] = []
        for fname, mt, raw in atts:
            safe = re.sub(r"[^a-zA-Z0-9._-]+", "_", fname)[:50] or "att"
            if mt == "application/pdf" or fname.lower().endswith(".pdf"):
                if raw.startswith(b"%PDF"):
                    p = TMPDIR / f"odoo_{eid}_{safe}"
                    if not p.suffix:
                        p = p.with_suffix(".pdf")
                    p.write_bytes(raw)
                    receipt_pdfs.append(p)
            elif mt.startswith("image/"):
                try:
                    pdf_bytes = L.image_to_pdf(raw)
                    p = TMPDIR / f"odoo_{eid}_{safe}.pdf"
                    p.write_bytes(pdf_bytes)
                    receipt_pdfs.append(p)
                except Exception as ex:
                    print(f"  ! image→pdf {fname}: {ex}", file=sys.stderr)

        detail = TMPDIR / f"detail_{i:03d}.pdf"
        build_detail(i, r, bool(receipt_pdfs), detail)
        pieces.append(detail)
        pieces.extend(receipt_pdfs)

    print(f"\nMerging {len(pieces)} PDFs → {out_path}", file=sys.stderr)
    writer = PdfWriter()
    for p in pieces:
        try:
            reader = PdfReader(str(p))
            for page in reader.pages:
                writer.add_page(page)
        except Exception as ex:
            print(f"  ! skip {p.name}: {ex}", file=sys.stderr)
    with open(out_path, "wb") as f:
        writer.write(f)
    sz = out_path.stat().st_size / 1024
    print(f"Wrote: {out_path} ({sz:.1f} KB · {len(pieces)} source pages)", file=sys.stderr)


def main(xlsx_path: str, pdf_path: str | None) -> None:
    xlsx = Path(xlsx_path).expanduser()
    if not xlsx.exists():
        raise SystemExit(f"XLSX not found: {xlsx}")

    rows = read_y_rows(xlsx)
    if not rows:
        raise SystemExit("No y-marked rows in spreadsheet — nothing to do.")
    print(f"{len(rows)} y-marked expenses\n", file=sys.stderr)

    # Period label from xlsx filename if it follows the collect.py convention
    m = re.search(r"(\d{4}-\d{2}-\d{2}).*?(\d{4}-\d{2}-\d{2})", xlsx.name)
    if m:
        period_label = f"{m.group(1)} to {m.group(2)}"
        period_slug = f"{m.group(1)}_to_{m.group(2)}"
    else:
        period_label = f"{rows[0]['date']} to {rows[-1]['date']}"
        period_slug = f"{rows[0]['date']}_to_{rows[-1]['date']}"

    if not pdf_path:
        pdf_path = f"/Users/davidorban/Downloads/Futuroid_Expenses_Report_{period_slug}.pdf"
    out_pdf = Path(pdf_path).expanduser()

    print("Auth Gmail...", file=sys.stderr)
    token = L.gmail_token()

    print("Pushing to Odoo + attaching receipts...", file=sys.stderr)
    push_to_odoo(rows, token)

    print("\nBuilding merged PDF...", file=sys.stderr)
    build_pdf(rows, period_label, out_pdf)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: finalize.py <xlsx_path> [pdf_out]", file=sys.stderr)
        sys.exit(2)
    main(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
    L.safe_exit(0)
